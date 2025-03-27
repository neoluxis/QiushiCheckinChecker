import sys
import os
import re
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QPushButton, QLabel,
    QFileDialog, QSlider, QHBoxLayout, QTableWidget, QTableWidgetItem
)
from PyQt6.QtCore import Qt


# === 工具函数 ===
def find_time(text):
    matches = re.findall(r"\b\d{2}:\d{2}\b(?![^\(]*\))", str(text))
    return matches


class MyTime:
    def __init__(self, time_str):
        self.h, self.m = map(int, time_str.split(":"))

    def __sub__(self, other):
        h1, m1 = self.h, self.m
        h2, m2 = other.h, other.m
        if h1 < h2:
            h1 += 24
        mins = (h1 - h2) * 60 + (m1 - m2)
        return MyTime(f"{mins // 60}:{mins % 60}")

    def __gt__(self, other):
        return self.total_minutes() > other.total_minutes()

    def __lt__(self, other):
        return self.total_minutes() < other.total_minutes()


    def total_minutes(self):
        return self.h * 60 + self.m

    def __str__(self):
        return f"{self.h:02d}:{self.m:02d}"


def process_xlsx(file_path, workday_cols, weekend_cols, name_col, weekdays=3, weekend=8, min_punch_time=4, weekend_full_time=12):
    df = pd.read_excel(file_path, header=None)
    personnes = df.iloc[4:]

    def calculate_time_diffs(cols):
        data = []
        for row in personnes.index:
            times = []
            for col in cols:
                ts = find_time(df.iloc[row, col])
                if len(ts) < 2:
                    times.append(MyTime('00:00'))
                    continue
                time2 = MyTime(ts[-1])
                time1 = MyTime(ts[0])
                times.append(time2 - time1)
            data.append(times)
        return data

    workday_times = calculate_time_diffs(workday_cols)
    weekend_times = calculate_time_diffs(weekend_cols)

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    red_fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
    green_fill = PatternFill(start_color="66FF66", end_color="66FF66", fill_type="solid")

    name_col_index = ord(name_col) - ord('A')
    selected_columns = [name_col_index] + workday_cols + weekend_cols  # 姓名列 + 工作日列 + 周末列

    # 处理每一行数据
    for i in range(len(personnes)):
        work_days = sum(1 for t in workday_times[i] if t.total_minutes() >= min_punch_time * 60)
        weekend_hours = sum(t.total_minutes() for t in weekend_times[i]) / 60

        if work_days < weekdays or weekend_hours < weekend:
            for col in selected_columns:
                ws.cell(row=i + 5, column=col + 1).fill = red_fill
        elif work_days >= len(workday_cols) and weekend_hours >= weekend_full_time:
            for col in selected_columns:
                ws.cell(row=i + 5, column=col + 1).fill = green_fill

    # 仅保存被标红或标绿的列
    for col in range(ws.max_column, 0, -1):
        if all(ws.cell(row=row, column=col).fill == PatternFill() for row in range(5, len(personnes) + 5)):
            ws.delete_cols(col)

    # 保存标注颜色后的文件
    output_file = f"output_{os.path.basename(file_path)}"
    wb.save(output_file)
    print(f"处理完成，结果已保存为 {output_file}")


# === PyQt6 GUI ===
class AttendanceApp(QWidget):
    def __init__(self):
        super().__init__()
        self.file_path = None
        self.workday_cols = []
        self.weekend_cols = []
        self.name_col = None
        self.selecting_mode = None
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("考勤分析工具")
        self.setGeometry(100, 100, 600, 600)
        layout = QVBoxLayout()

        self.file_label = QLabel("拖放或选择Excel文件")
        layout.addWidget(self.file_label)

        self.file_button = QPushButton("选择文件")
        self.file_button.clicked.connect(self.select_file)
        layout.addWidget(self.file_button)

        self.table = QTableWidget()
        self.table.cellClicked.connect(self.select_column)
        layout.addWidget(self.table)

        # 选择姓名列
        self.name_layout = QHBoxLayout()
        self.name_label = QLabel("姓名列: 无")
        self.name_button = QPushButton("选择姓名列")
        self.name_button.clicked.connect(self.set_selecting_mode_name)
        self.name_layout.addWidget(self.name_label)
        self.name_layout.addWidget(self.name_button)
        layout.addLayout(self.name_layout)

        workday_layout = QHBoxLayout()
        self.workday_label = QLabel("工作日列: 无")
        self.workday_button = QPushButton("选择工作日列")
        self.workday_button.clicked.connect(lambda: self.set_selecting_mode("workday"))
        workday_layout.addWidget(self.workday_label)
        workday_layout.addWidget(self.workday_button)
        layout.addLayout(workday_layout)

        weekend_layout = QHBoxLayout()
        self.weekend_label = QLabel("周末列: 无")
        self.weekend_button = QPushButton("选择周末列")
        self.weekend_button.clicked.connect(lambda: self.set_selecting_mode("weekend"))
        weekend_layout.addWidget(self.weekend_label)
        weekend_layout.addWidget(self.weekend_button)
        layout.addLayout(weekend_layout)

        self.weekdays_label = QLabel("最低工作日天数: 3")
        layout.addWidget(self.weekdays_label)
        self.weekdays_slider = QSlider(Qt.Orientation.Horizontal)
        self.weekdays_slider.setMinimum(1)
        self.weekdays_slider.setMaximum(5)
        self.weekdays_slider.setValue(3)
        self.weekdays_slider.setTickInterval(1)
        self.weekdays_slider.setSingleStep(1)
        self.weekdays_slider.valueChanged.connect(self.update_weekdays_label)
        layout.addWidget(self.weekdays_slider)

        self.weekend_label_slider = QLabel("最低周末时长 (小时): 8")
        layout.addWidget(self.weekend_label_slider)
        self.weekend_slider = QSlider(Qt.Orientation.Horizontal)
        self.weekend_slider.setMinimum(0)
        self.weekend_slider.setMaximum(32)
        self.weekend_slider.setValue(8)
        self.weekend_slider.setTickInterval(1)
        self.weekend_slider.setSingleStep(1)
        self.weekend_slider.valueChanged.connect(self.update_weekend_label)
        layout.addWidget(self.weekend_slider)

        self.min_punch_label = QLabel("单日最小打卡时间 (小时): 4")
        layout.addWidget(self.min_punch_label)
        self.min_punch_slider = QSlider(Qt.Orientation.Horizontal)
        self.min_punch_slider.setMinimum(0)
        self.min_punch_slider.setMaximum(8)
        self.min_punch_slider.setValue(1)
        self.min_punch_slider.setTickInterval(1)
        self.min_punch_slider.setSingleStep(1)
        self.min_punch_slider.valueChanged.connect(self.update_min_punch_label)
        layout.addWidget(self.min_punch_slider)

        self.full_time_label = QLabel("周末满勤时长 (小时): 12")
        layout.addWidget(self.full_time_label)
        self.full_time_slider = QSlider(Qt.Orientation.Horizontal)
        self.full_time_slider.setMinimum(0)
        self.full_time_slider.setMaximum(32)
        self.full_time_slider.setValue(10)
        self.full_time_slider.setTickInterval(1)
        self.full_time_slider.setSingleStep(1)
        self.full_time_slider.valueChanged.connect(self.update_full_time_label)
        layout.addWidget(self.full_time_slider)

        self.process_button = QPushButton("开始处理")
        self.process_button.clicked.connect(self.run_processing)
        layout.addWidget(self.process_button)

        self.quit_button = QPushButton("退出")
        self.quit_button.clicked.connect(self.close)
        layout.addWidget(self.quit_button)

        self.setLayout(layout)

        # Enable drag and drop
        self.setAcceptDrops(True)

    def select_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "选择Excel文件", "", "Excel Files (*.xlsx)")
        if file_path:
            self.file_label.setText(file_path)
            self.file_path = file_path
            self.load_preview()

    def load_preview(self):
        df = pd.read_excel(self.file_path, header=None)
        self.table.setRowCount(min(10, len(df)))
        self.table.setColumnCount(len(df.columns))
        self.table.setHorizontalHeaderLabels([chr(65 + i) for i in range(len(df.columns))])

        for i in range(min(10, len(df))):
            for j in range(len(df.columns)):
                self.table.setItem(i, j, QTableWidgetItem(str(df.iloc[i, j])))

    def set_selecting_mode(self, mode):
        self.selecting_mode = mode

    def set_selecting_mode_name(self):
        self.selecting_mode = "name"

    def select_column(self, row, col):
        if self.selecting_mode == "workday":
            if col not in self.workday_cols:
                self.workday_cols.append(col)
                self.workday_label.setText(f"工作日列: {', '.join(chr(65 + c) for c in self.workday_cols)}")
        elif self.selecting_mode == "weekend":
            if col not in self.weekend_cols:
                self.weekend_cols.append(col)
                self.weekend_label.setText(f"周末列: {', '.join(chr(65 + c) for c in self.weekend_cols)}")
        elif self.selecting_mode == "name":
            self.name_col = chr(65 + col)
            self.name_label.setText(f"姓名列: {self.name_col}")

    def update_weekdays_label(self, value):
        self.weekdays_label.setText(f"最低工作日天数: {value}")

    def update_weekend_label(self, value):
        self.weekend_label_slider.setText(f"最低周末时长 (小时): {value}")

    def update_min_punch_label(self, value):
        self.min_punch_label.setText(f"单日最小打卡时间 (小时): {value}")

    def update_full_time_label(self, value):
        self.full_time_label.setText(f"周末满勤时长 (小时): {value}")

    def run_processing(self):
        if not self.file_path or not self.workday_cols or not self.weekend_cols or not self.name_col:
            self.file_label.setText("请先选择文件和列！")
            return

        weekdays = self.weekdays_slider.value()
        weekend = self.weekend_slider.value()
        min_punch_time = self.min_punch_slider.value()
        weekend_full_time = self.full_time_slider.value()

        process_xlsx(self.file_path, self.workday_cols, self.weekend_cols, self.name_col, weekdays, weekend, min_punch_time, weekend_full_time)

        self.file_label.setText("✅ 处理完成！文件已保存")

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event):
        file_path = event.mimeData().urls()[0].toLocalFile()
        if file_path.endswith(".xlsx"):
            self.file_label.setText(file_path)
            self.file_path = file_path
            self.load_preview()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = AttendanceApp()
    window.show()
    sys.exit(app.exec())
